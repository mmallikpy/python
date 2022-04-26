import os
import socket
from openpyxl import load_workbook  # For Excel.

# https://www.youtube.com/watch?v=TluyP4n6n-U
# IP Series List
# IP list using ping
# Find FQDN
# Create Excel and wirite the output
li_subnet = ["172.16.7"]
li_ip = []
excel_sheet = load_workbook("ip_fqdn.xlsx")

for subnet in li_subnet:
    sheetNames = excel_sheet.sheetnames
    if subnet not in sheetNames:
        print(subnet)
        excel_sheet.create_sheet(subnet)
        excel_sheet.save("ip_fqdn.xlsx")

dicForInsert = {}
for subnetFP in li_subnet:
    for ip in range(11, 15):
        output = os.system(f"ping {subnetFP}.{ip} -c 2")
        if output == 0:
            fqdn = socket.getfqdn(f"{subnetFP}.{ip}")
            print("----------", f"{subnetFP}.{ip}")
            print("------", fqdn)
            dicForInsert[f"{subnetFP}.{ip}"] = fqdn
            excel_sheet1 = excel_sheet.active
            excel_sheet1.append(dicForInsert)[excel_sheet("172.16.7")]
            excel_sheet.save("ip_fqdn.xlsx")
